import os
import openpyxl

def add_to_file(object_key, input):
    with open(object_key, 'a') as file:
        file.write('\n' + input + '\n')

def remove_file(object_key):
    os.remove(object_key)

def add_line_to_xlsx(object_key, input):
    workbook = openpyxl.load_workbook(object_key)
    sheet = workbook.active
    new_row = input
    sheet.append(new_row)
    workbook.save(object_key)

def read_from_xlsx(object_key):
    workbook = openpyxl.load_workbook(object_key)
    sheet = workbook.active
    for row in sheet.iter_rows(values_only=True):
        print(row)